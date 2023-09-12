import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook
from datetime import datetime
from tkcalendar import Calendar

def add_data_to_excel():
    customer = customer_combobox.get()
    site = site_combobox.get()
    date = date_cal.get_date()
    net = net_entry.get()
    version = version_entry.get()
    total_event = total_event_entry.get()
    false_event = false_event_entry.get()
    version_bugs = version_bugs_entry.get()
    unique_id = unique_id_entry.get()
    excel_file_path = "O:\\QA\\live.xlsx"

    try:
        workbook = load_workbook(excel_file_path)
        if customer in workbook.sheetnames:
            sheet = workbook[customer]
        else:
            sheet = workbook.create_sheet(customer)

        row = [date, site, net, version, total_event, false_event, version_bugs, unique_id]
        sheet.append(row)

        # Add the data to the "follow up" sheet
        follow_up_sheet = workbook["follow up"]

        # Check if the date already exists in the "follow up" sheet
        date_column = follow_up_sheet["A"]
        dates = [cell.value for cell in date_column[1:]]  # Skip the header
        if date in dates:
            row_index = dates.index(date) + 2  # Add 2 to skip the header and convert to 1-based index
        else:
            follow_up_sheet.insert_rows(2)  # Insert a new row at row 2 (adjust as needed)
            follow_up_sheet.cell(row=2, column=1, value=date)
            row_index = 2

        # Mark the site column with a "V"
        site_index = sites.index(site) + 2  # Add 2 to skip the header and convert to 1-based index
        follow_up_sheet.cell(row=row_index, column=site_index, value="V")

        workbook.save(excel_file_path)
        result_label.config(text="Data added to Excel successfully.")
    except Exception as e:
        result_label.config(text=f"Error adding data to Excel: {str(e)}")

# Create the main window
root = tk.Tk()
root.title("Excel Data Entry")

# Create and configure the frame
frame = ttk.Frame(root, padding=10)
frame.grid(column=0, row=0, sticky=(tk.W, tk.E, tk.N, tk.S))
frame.columnconfigure(1, weight=1)

# Customer selection
customer_label = ttk.Label(frame, text="Customer:")
customer_label.grid(column=0, row=0, sticky=tk.W)
customer_combobox = ttk.Combobox(frame, values=["cemex", "Shafir", "hidenberg"])
customer_combobox.grid(column=1, row=0, sticky=tk.W)

# Site selection
site_label = ttk.Label(frame, text="Site:")
site_label.grid(column=0, row=1, sticky=tk.W)
sites = ["adhalom","Gdansk", "Gdynia" , "golani","mevocarmel", "modiim", "yvnehe" , "roller", "Etziyona", "London1627","London1656"]
site_combobox = ttk.Combobox(frame, values=[])
site_combobox.grid(column=1, row=1, sticky=tk.W)

# Date entry
date_label = ttk.Label(frame, text="Date:")
date_label.grid(column=0, row=2, sticky=tk.W)
date_cal = Calendar(frame)
date_cal.grid(column=1, row=2, sticky=tk.W)

# Other fields (net, version, total event, false event, version bugs, unique id)
net_label = ttk.Label(frame, text="Net:")
net_label.grid(column=0, row=3, sticky=tk.W)
net_entry = ttk.Entry(frame)
net_entry.grid(column=1, row=3, sticky=tk.W)

version_label = ttk.Label(frame, text="Version:")
version_label.grid(column=0, row=4, sticky=tk.W)
version_entry = ttk.Entry(frame)
version_entry.grid(column=1, row=4, sticky=tk.W)

total_event_label = ttk.Label(frame, text="Total Event:")
total_event_label.grid(column=0, row=5, sticky=tk.W)
total_event_entry = ttk.Entry(frame)
total_event_entry.grid(column=1, row=5, sticky=tk.W)

false_event_label = ttk.Label(frame, text="False Event:")
false_event_label.grid(column=0, row=6, sticky=tk.W)
false_event_entry = ttk.Entry(frame)
false_event_entry.grid(column=1, row=6, sticky=tk.W)

version_bugs_label = ttk.Label(frame, text="Version Bugs:")
version_bugs_label.grid(column=0, row=7, sticky=tk.W)
version_bugs_entry = ttk.Entry(frame)
version_bugs_entry.grid(column=1, row=7, sticky=tk.W)

unique_id_label = ttk.Label(frame, text="Unique ID:")
unique_id_label.grid(column=0, row=8, sticky=tk.W)
unique_id_entry = ttk.Entry(frame)
unique_id_entry.grid(column=1, row=8, sticky=tk.W)

# Excel file path entry
#excel_file_path_label = ttk.Label(frame, text="Excel File Path:")
#excel_file_path_label.grid(column=0, row=9, sticky=tk.W)
#excel_file_path_entry = ttk.Entry(frame)
#excel_file_path_entry.grid(column=1, row=9, sticky=(tk.W, tk.E))
#excel_file_path_entry.insert(0, "O:\\QA\\live.xlsx")

# Add data button
add_data_button = ttk.Button(frame, text="Add to Excel", command=add_data_to_excel)
add_data_button.grid(column=1, row=10, sticky=tk.E)

# Result label
result_label = ttk.Label(frame, text="")
result_label.grid(column=0, row=11, columnspan=2)

# Update site choices based on customer selection
def update_site_choices(event):
    selected_customer = customer_combobox.get()
    if selected_customer == "Shafir":
        site_combobox['values'] = ["roller", "Etziyona"]
    elif selected_customer == "cemex":
        site_combobox['values'] = ["adhalom","Gdansk", "Gdynia" , "golani","mevocarmel", "modiim", "yvnehe"]
    elif selected_customer == "hidenberg":
        site_combobox['values'] = ["London"]
    else:
        site_combobox['values'] = []

customer_combobox.bind("<<ComboboxSelected>>", update_site_choices)

root.mainloop()
